"""Parser for JBMO 2025 results.

Three raw files in data/jbmo/raw/2025/:

- ``jbmo_2025_scores.xlsx`` (manually supplied): per-problem scores keyed
  by ``"<COUNTRY> <NUM>"`` rows like ``ALB 1`` … ``UZB 6``. Columns are
  ``[code, P1, P2, P3, P4, Total]`` with separator/header rows between
  countries.
- ``jbmo_2025_participants.html`` (downloaded from
  web.archive.org): provides the ``(country, num) -> name`` mapping. The
  participants are rendered as ``<div class="portfolio-item filter-XXX">``
  blocks where each block contains ``"XXX N Given Family"`` text. Leader
  and Deputy entries are skipped.
- ``jbmo_2025_raw.html`` (Excel-exported HTML): used only as the source
  of truth for award names — provides ``(total -> award)`` lookups.

Names are in "Given Surname" order. Country codes are 3-letter (or with a
``-B`` suffix, e.g. ``MKD-B``).
"""

import re
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup

from ..models import ContestantResult
from .base import BaseParser

# Strip wrapping whitespace and collapse internal whitespace runs.
_WS_RE = re.compile(r"\s+")


def _clean(text: str) -> str:
    return _WS_RE.sub(" ", text).strip()


def _split_name(full_name: str) -> tuple[str | None, str | None]:
    tokens = full_name.strip().split()
    if len(tokens) < 2:
        return None, None
    return " ".join(tokens[:-1]), tokens[-1]


# Matches a participant block label like "ALB 1 Eol Myshketa" or "MKD-B 3 …".
_PARTICIPANT_RE = re.compile(r"^(\S+)\s+(\d+)\s+(.+)$")


def _parse_participants(participants_file: Path) -> dict[tuple[str, int], str]:
    """Return a {(country_code, num): full_name} mapping from the participants page."""
    soup = BeautifulSoup(participants_file.read_text(encoding="utf-8"), "html.parser")
    mapping: dict[tuple[str, int], str] = {}

    for item in soup.select("div.portfolio-item"):
        text = _clean(item.get_text(" ", strip=True))
        match = _PARTICIPANT_RE.match(text)
        if not match:
            continue
        code, num_str, name = match.groups()
        # Skip non-contestant entries (Deputy, Leader, Observer, etc.).
        if not num_str.isdigit():
            continue
        mapping[(code.upper(), int(num_str))] = name

    return mapping


def _parse_scores(xlsx_file: Path) -> list[tuple[str, int, list[int | None], int]]:
    """Read the score sheet and return [(country, num, [p1..p4], total), ...]."""
    workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[tuple[str, int, list[int | None], int]] = []
    label_re = re.compile(r"^([A-Za-z][A-Za-z\-]*)\s+(\d+)$")

    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        match = label_re.match(label)
        if not match:
            # Header row ("None / 1 / 2 / 3 / 4 / Total") or anything else.
            continue
        country = match.group(1).upper()
        num = int(match.group(2))

        if row[5] is None:
            continue

        scores = [None if v is None else int(v) for v in row[1:5]]
        total = int(row[5])
        rows.append((country, num, scores, total))

    return rows


def _parse_medal_cutoffs(awards_file: Path) -> dict[str, int]:
    """Return min total score for each medal tier from the totals-only HTML.

    HMs are excluded — they are not strictly score-based (JBMO awards an
    HM to non-medalists with at least one full-marks problem score).
    """
    soup = BeautifulSoup(awards_file.read_text(encoding="utf-8"), "html.parser")
    table = soup.find("table")
    if not table:
        raise ValueError("Could not find results table in JBMO 2025 awards HTML")

    by_award: dict[str, list[int]] = {"Gold": [], "Silver": [], "Bronze": []}
    for tr in table.find_all("tr"):
        cells = [_clean(c.get_text(" ", strip=True)) for c in tr.find_all(["td", "th"])]
        if len(cells) < 5 or not cells[3].isdigit():
            continue
        total = int(cells[3])
        award = cells[4]
        for tier in by_award:
            if tier in award:
                by_award[tier].append(total)

    return {tier: min(totals) for tier, totals in by_award.items() if totals}


def _award_for(total: int, scores: list[int | None], cutoffs: dict[str, int]) -> str | None:
    """Determine award given total, per-problem scores, and medal cutoffs.

    Honourable Mention is awarded to non-medalists who got full marks (10)
    on at least one problem.
    """
    if "Gold" in cutoffs and total >= cutoffs["Gold"]:
        return "Gold"
    if "Silver" in cutoffs and total >= cutoffs["Silver"]:
        return "Silver"
    if "Bronze" in cutoffs and total >= cutoffs["Bronze"]:
        return "Bronze"
    if any(s == 10 for s in scores if s is not None):
        return "Honourable Mention"
    return None


class Parser2025(BaseParser):
    """Parser for JBMO 2025 (XLSX scores + participants HTML names + awards HTML)."""

    def parse(self, raw_file: Path) -> list[ContestantResult]:
        # ``raw_file`` is the participants HTML returned by the downloader.
        # The XLSX and awards HTML live alongside it.
        raw_dir = raw_file.parent
        participants_file = raw_file
        xlsx_file = raw_dir / "jbmo_2025_scores.xlsx"
        awards_file = raw_dir / "jbmo_2025_raw.html"

        if not xlsx_file.exists():
            raise FileNotFoundError(
                f"JBMO 2025 score sheet not found: {xlsx_file}. Place the XLSX "
                "scores file there before parsing."
            )

        names = _parse_participants(participants_file)
        score_rows = _parse_scores(xlsx_file)
        cutoffs = _parse_medal_cutoffs(awards_file) if awards_file.exists() else {}

        results: list[ContestantResult] = []
        for country, num, scores, total in score_rows:
            name = names.get((country, num))
            if name is None:
                raise ValueError(f"No name found in participants page for {country} {num}")

            given_name, family_name = _split_name(name)
            award = self.normalize_award(_award_for(total, scores, cutoffs))

            results.append(
                ContestantResult(
                    name=name,
                    country=country,
                    problem_scores=scores,
                    total=total,
                    rank=None,  # Computed below.
                    award=award,
                    given_name=given_name,
                    family_name=family_name,
                )
            )

        self.compute_ranks(results)
        return results
